"""Rutas compartidas: index, status CDC, visor de datos CDC, truncate CDC, Excel export."""
import os
from datetime import datetime
from flask import Blueprint, render_template, request, jsonify
import time as _time
import urllib.request as _urlreq
from config import DATABASES, get_connection
from db import get_kafka_outbox_offset, get_schemas_and_tables
from modules import ALL_MODULES

common_bp = Blueprint('common', __name__)


@common_bp.route("/")
def index():
    return render_template("index.html")


@common_bp.route("/api/status")
def get_status():
    """Estado de las colas CDC - optimizado con NOLOCK y query unica"""
    status = {}
    try:
        for db_key, db_name in DATABASES.items():
            conn = get_connection(db_key)
            cursor = conn.cursor()
            try:
                cursor.execute("""
                    SELECT
                        (SELECT COUNT_BIG(*) FROM dbo.cdc_outbox WITH (NOLOCK) WHERE aggregate_type != '_warmup') AS outbox,
                        (SELECT COUNT_BIG(*) FROM dbo.cdc_inbox WITH (NOLOCK) WHERE processed = 0) AS inbox_pending,
                        (SELECT COUNT_BIG(*) FROM dbo.cdc_inbox WITH (NOLOCK) WHERE processed = 1) AS inbox_processed,
                        (SELECT COUNT_BIG(*) FROM dbo.cdc_inbox_errors WITH (NOLOCK)) AS errors
                """)
                row = cursor.fetchone()
                status[f"{db_key}_outbox"] = row[0]
                status[f"{db_key}_inbox_pending"] = row[1]
                status[f"{db_key}_inbox_processed"] = row[2]
                status[f"{db_key}_errors"] = row[3]
            except:
                status[f"{db_key}_outbox"] = "N/A"
                status[f"{db_key}_inbox_pending"] = "N/A"
                status[f"{db_key}_inbox_processed"] = "N/A"
                status[f"{db_key}_errors"] = "N/A"
            conn.close()
    except Exception as e:
        status["error"] = str(e)

    return jsonify(status)


@common_bp.route("/api/cdc-pending-summary/<db_key>")
def get_pending_summary(db_key):
    """Diagnostico de registros pendientes en cdc_inbox"""
    if db_key not in DATABASES:
        return jsonify({"error": "DB invalida"}), 400

    try:
        conn = get_connection(db_key)
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM dbo.cdc_inbox WHERE processed = 0")
        total = cursor.fetchone()[0]

        if total == 0:
            conn.close()
            return jsonify({"total": 0, "groups": [], "oldest": None, "newest": None})

        cursor.execute("""
            SELECT aggregate_type, event_type, source_table,
                   COUNT(*) as cnt,
                   MIN(created_at) as oldest,
                   MAX(created_at) as newest
            FROM dbo.cdc_inbox WHERE processed = 0
            GROUP BY aggregate_type, event_type, source_table
            ORDER BY cnt DESC
        """)
        groups = []
        for r in cursor.fetchall():
            groups.append({
                "aggregate_type": r.aggregate_type,
                "event_type": r.event_type,
                "source_table": r.source_table or "N/A",
                "count": r.cnt,
                "oldest": str(r.oldest),
                "newest": str(r.newest)
            })

        cursor.execute("""
            SELECT MIN(created_at), MAX(created_at),
                   DATEDIFF(MINUTE, MAX(created_at), GETDATE()) as minutes_ago
            FROM dbo.cdc_inbox WHERE processed = 0
        """)
        row = cursor.fetchone()
        oldest = str(row[0]) if row[0] else None
        newest = str(row[1]) if row[1] else None
        minutes_stuck = row[2] if row[2] else 0

        conn.close()
        return jsonify({
            "total": total, "groups": groups,
            "oldest": oldest, "newest": newest,
            "minutes_stuck": minutes_stuck
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@common_bp.route("/api/cdc-data/<db_key>/<table_name>")
def get_cdc_data(db_key, table_name):
    """Retorna datos de cualquier tabla CDC con paginacion"""
    allowed = ['cdc_inbox', 'cdc_outbox', 'cdc_inbox_errors']
    if db_key not in DATABASES:
        return jsonify({"error": "DB invalida"}), 400
    if table_name not in allowed:
        return jsonify({"error": "Tabla no permitida"}), 400

    filter_type = request.args.get("filter")
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 100))
    offset = (page - 1) * per_page

    try:
        conn = get_connection(db_key)
        cursor = conn.cursor()

        cursor.execute(f"""
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = '{table_name}'
            ORDER BY ORDINAL_POSITION
        """)
        col_names = [row.COLUMN_NAME for row in cursor.fetchall()]

        if not col_names:
            conn.close()
            return jsonify({"columns": [], "rows": [], "total": 0, "page": page, "pages": 0})

        cols_sql = ", ".join([f"[{c}]" for c in col_names])

        where = ""
        if table_name == 'cdc_inbox' and filter_type == 'pending':
            where = "WHERE processed = 0"
        elif table_name == 'cdc_inbox' and filter_type == 'processed':
            where = "WHERE processed = 1"

        cursor.execute(f"SELECT COUNT(*) FROM dbo.[{table_name}] {where}")
        total = cursor.fetchone()[0]
        total_pages = (total + per_page - 1) // per_page if total > 0 else 1

        cursor.execute(f"""
            SELECT {cols_sql} FROM dbo.[{table_name}] {where}
            ORDER BY 1 DESC
            OFFSET {offset} ROWS FETCH NEXT {per_page} ROWS ONLY
        """)
        rows = []
        for row in cursor.fetchall():
            rows.append([str(v) if v is not None else "" for v in row])

        conn.close()
        return jsonify({
            "columns": col_names, "rows": rows, "total": total,
            "page": page, "pages": total_pages, "per_page": per_page
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@common_bp.route("/api/truncate-cdc", methods=["POST"])
def truncate_cdc():
    """Truncate a CDC table (inbox, outbox, inbox_errors)"""
    data = request.json
    db_key = data.get("db_key")
    table = data.get("table")

    allowed_tables = ['cdc_inbox', 'cdc_outbox', 'cdc_inbox_errors']
    if db_key not in DATABASES:
        return jsonify({"error": "DB invalida"}), 400
    if table not in allowed_tables:
        return jsonify({"error": "Tabla no permitida"}), 400

    try:
        conn = get_connection(db_key)
        cursor = conn.cursor()

        cursor.execute(f"SELECT COUNT(*) FROM dbo.[{table}]")
        before = cursor.fetchone()[0]

        # Si es outbox, obtener offset de Kafka ANTES de truncar
        kafka_offset = None
        if table == "cdc_outbox":
            kafka_offset = get_kafka_outbox_offset(db_key)

        try:
            cursor.execute(f"TRUNCATE TABLE dbo.[{table}]")
            conn.commit()
            method = "TRUNCATE"
        except Exception:
            conn.rollback()
            cursor.execute(f"DELETE FROM dbo.[{table}]")
            conn.commit()
            method = "DELETE"

        # Reseedear identity del outbox para que Kafka siga leyendo
        if table == "cdc_outbox" and kafka_offset:
            cursor.execute(f"DBCC CHECKIDENT ('dbo.cdc_outbox', RESEED, {int(kafka_offset)})")
            conn.commit()
            method += f" + RESEED identity a {kafka_offset}"

        conn.close()
        return jsonify({"table": f"{db_key}.dbo.{table}", "deleted": before, "method": method})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@common_bp.route("/api/truncate-module", methods=["POST"])
def truncate_module():
    """Trunca todas las tablas TYPE de un modulo en newcore + sus tablas originales en dbXX"""
    data = request.json
    module_name = data.get("module")

    if module_name not in ALL_MODULES:
        return jsonify({"error": f"Modulo '{module_name}' no encontrado"}), 400

    mod = ALL_MODULES[module_name]
    results = []

    # 1. Truncar tablas TYPE en fcme_newcore
    try:
        conn = get_connection("newcore")
        cursor = conn.cursor()

        # Obtener tablas del schema/modulo en newcore
        type_tables = list(mod.get("NEWCORE_TO_FINAL", {}).keys())

        for table in type_tables:
            try:
                cursor.execute(f"SELECT COUNT(*) FROM [{module_name}].[{table}]")
                count = cursor.fetchone()[0]
                if count == 0:
                    results.append({"table": f"newcore.{module_name}.{table}", "deleted": 0, "status": "vacia"})
                    continue

                try:
                    cursor.execute(f"DISABLE TRIGGER ALL ON [{module_name}].[{table}]")
                    conn.commit()
                except:
                    pass

                try:
                    cursor.execute(f"TRUNCATE TABLE [{module_name}].[{table}]")
                    conn.commit()
                    method = "TRUNCATE"
                except:
                    conn.rollback()
                    cursor.execute(f"DELETE FROM [{module_name}].[{table}]")
                    conn.commit()
                    method = "DELETE"

                try:
                    cursor.execute(f"ENABLE TRIGGER ALL ON [{module_name}].[{table}]")
                    conn.commit()
                except:
                    pass

                results.append({"table": f"newcore.{module_name}.{table}", "deleted": count, "status": method})
            except Exception as e:
                results.append({"table": f"newcore.{module_name}.{table}", "deleted": 0, "status": f"error: {str(e)[:80]}"})

        conn.close()
    except Exception as e:
        results.append({"table": "newcore (conexion)", "deleted": 0, "status": f"error: {str(e)[:80]}"})

    # 2. Truncar tablas originales en dbXX
    original_db = mod.get("ORIGINAL_DB")
    final_tables = mod.get("NEWCORE_TO_FINAL", {})

    # Recopilar tablas unicas de dbXX
    db_tables = {}  # {db_name: set(table_names)}
    for type_table, (db_name, orig_table) in final_tables.items():
        if db_name not in db_tables:
            db_tables[db_name] = set()
        db_tables[db_name].add(orig_table)

    for db_name, tables in db_tables.items():
        try:
            conn = get_connection(db_name)
            cursor = conn.cursor()

            for table in sorted(tables):
                try:
                    cursor.execute(f"SELECT COUNT(*) FROM dbo.[{table}]")
                    count = cursor.fetchone()[0]
                    if count == 0:
                        results.append({"table": f"{db_name}.dbo.{table}", "deleted": 0, "status": "vacia"})
                        continue

                    try:
                        cursor.execute(f"DISABLE TRIGGER ALL ON dbo.[{table}]")
                        conn.commit()
                    except:
                        pass

                    try:
                        cursor.execute(f"TRUNCATE TABLE dbo.[{table}]")
                        conn.commit()
                        method = "TRUNCATE"
                    except:
                        conn.rollback()
                        cursor.execute(f"DELETE FROM dbo.[{table}]")
                        conn.commit()
                        method = "DELETE"

                    try:
                        cursor.execute(f"ENABLE TRIGGER ALL ON dbo.[{table}]")
                        conn.commit()
                    except:
                        pass

                    results.append({"table": f"{db_name}.dbo.{table}", "deleted": count, "status": method})
                except Exception as e:
                    results.append({"table": f"{db_name}.dbo.{table}", "deleted": 0, "status": f"error: {str(e)[:80]}"})

            conn.close()
        except Exception as e:
            results.append({"table": f"{db_name} (conexion)", "deleted": 0, "status": f"error: {str(e)[:80]}"})

    total_deleted = sum(r["deleted"] for r in results)
    return jsonify({"module": module_name, "total_deleted": total_deleted, "details": results})


@common_bp.route("/api/generate-excel")
def generate_excel():
    """Regenera el Excel completo desde el CSV"""
    try:
        reportes_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reportes")
        _update_excel(reportes_dir, [], [], {})
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@common_bp.route("/api/export-result", methods=["POST"])
def export_result():
    """Agrega resultado al CSV (git) y genera Excel local con formato"""
    try:
        import csv

        data = request.json
        if not data or "direction" not in data:
            return jsonify({"error": "Sin datos"}), 400

        reportes_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reportes")
        os.makedirs(reportes_dir, exist_ok=True)

        headers = [
            "Fecha/Hora", "Direccion", "Modulo", "Tabla", "Tabla Destino",
            "Cantidad", "Insertados", "Tiempo (seg)", "Rows/seg",
            "Count Antes", "Count Despues", "Neto",
            "Outbox", "Triggers", "Errores"
        ]

        row_data = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            data.get("direction", "").replace("_", " ").title(),
            data.get("schema", ""),
            data.get("table", ""),
            data.get("dest_table", ""),
            data.get("quantity", 0),
            data.get("total_inserted", 0),
            data.get("total_time_sec", 0),
            data.get("avg_rows_per_sec", 0),
            data.get("count_before", 0),
            data.get("count_after", 0),
            data.get("net_inserted", 0),
            data.get("outbox_generated", 0),
            "OFF" if data.get("triggers_disabled") else "ON",
            data.get("total_errors", 0),
        ]

        # 1. CSV (se sube al git)
        csv_path = os.path.join(reportes_dir, "resultados_cdc.csv")
        csv_exists = os.path.exists(csv_path)
        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not csv_exists:
                writer.writerow(headers)
            writer.writerow(row_data)

        # 2. Excel local con formato (no se sube al git)
        _update_excel(reportes_dir, headers, row_data, data)

        return jsonify({"file": "resultados_cdc.csv"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _update_excel(reportes_dir, headers, row_data, data):
    """Reconstruye el Excel completo desde el CSV con colores y formato"""
    try:
        import csv
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        csv_path = os.path.join(reportes_dir, "resultados_cdc.csv")
        xlsx_path = os.path.join(reportes_dir, "resultados_cdc.xlsx")

        if not os.path.exists(csv_path):
            return

        # Leer todo el CSV
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            all_rows = list(reader)

        if len(all_rows) < 2:
            return

        csv_headers = all_rows[0]
        csv_data = sorted(all_rows[1:], key=lambda r: r[0] if r else '', reverse=True)

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        err_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center = Alignment(horizontal='center', vertical='center')
        widths = [20, 24, 22, 32, 32, 14, 14, 14, 14, 14, 16, 12, 12, 12, 12]

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados CDC"

        # Titulo
        last_col = get_column_letter(len(csv_headers))
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'].value = "CDC Bulk Tester - Registro de Resultados"
        ws['A1'].font = Font(bold=True, size=14, color="2F5496")
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'A2:{last_col}2')
        ws['A2'].value = f"Actualizado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=9, color="999999")
        ws['A2'].alignment = Alignment(horizontal='center')

        # Headers
        ws.row_dimensions[4].height = 30
        for col, h in enumerate(csv_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Anchos
        for i, w in enumerate(widths[:len(csv_headers)]):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        ws.freeze_panes = 'A5'
        ws.auto_filter.ref = f"A4:{last_col}4"

        # Indices de columnas clave
        def col_idx(name):
            try:
                return csv_headers.index(name)
            except ValueError:
                return -1

        idx_cantidad = col_idx("Cantidad")
        idx_insertados = col_idx("Insertados")
        idx_errores = col_idx("Errores")
        idx_outbox = col_idx("Outbox")
        idx_triggers = col_idx("Triggers")

        # Datos
        for row_idx, csv_row in enumerate(csv_data):
            excel_row = row_idx + 5
            is_alt = row_idx % 2 == 1

            for col, val in enumerate(csv_row, 1):
                # Intentar convertir numeros
                try:
                    if '.' in val:
                        val = float(val)
                    else:
                        val = int(val)
                except (ValueError, TypeError):
                    pass

                cell = ws.cell(row=excel_row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center
                cell.font = Font(size=10)
                if is_alt:
                    cell.fill = alt_fill

            # Colores condicionales
            try:
                cantidad = int(csv_row[idx_cantidad]) if idx_cantidad >= 0 and csv_row[idx_cantidad] else 0
                insertados = int(csv_row[idx_insertados]) if idx_insertados >= 0 and csv_row[idx_insertados] else 0
                errores = int(csv_row[idx_errores]) if idx_errores >= 0 and csv_row[idx_errores] else 0
                outbox = int(csv_row[idx_outbox]) if idx_outbox >= 0 and csv_row[idx_outbox] else 0
                triggers = csv_row[idx_triggers] if idx_triggers >= 0 else "ON"

                # Insertados
                if idx_insertados >= 0:
                    c = ws.cell(row=excel_row, column=idx_insertados + 1)
                    if insertados == cantidad and cantidad > 0:
                        c.fill = ok_fill
                        c.font = Font(size=10, bold=True, color="006100")
                    elif cantidad > 0:
                        c.fill = err_fill
                        c.font = Font(size=10, bold=True, color="9C0006")

                # Errores
                if idx_errores >= 0:
                    c = ws.cell(row=excel_row, column=idx_errores + 1)
                    if errores > 0:
                        c.fill = err_fill
                        c.font = Font(size=10, bold=True, color="9C0006")
                    else:
                        c.fill = ok_fill
                        c.font = Font(size=10, color="006100")

                # Outbox
                if idx_outbox >= 0:
                    c = ws.cell(row=excel_row, column=idx_outbox + 1)
                    if outbox > 0:
                        c.fill = ok_fill
                    elif triggers == "ON":
                        c.fill = warn_fill

                # Triggers
                if idx_triggers >= 0:
                    c = ws.cell(row=excel_row, column=idx_triggers + 1)
                    if triggers == "OFF":
                        c.fill = warn_fill
            except:
                pass

        wb.save(xlsx_path)
    except ImportError:
        pass
    except:
        pass
